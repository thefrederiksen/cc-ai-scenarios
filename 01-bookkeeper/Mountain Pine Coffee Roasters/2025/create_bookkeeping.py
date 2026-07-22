"""
Create Bookkeeping 2025.xlsx from bookkeeping_spec.json
"""
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SPEC_PATH = os.path.join(SCRIPT_DIR, "bookkeeping_spec.json")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "Bookkeeping 2025.xlsx")

# Boardroom theme colors
THEME = {
    "header": {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    },
    "subheader": {
        "font": Font(bold=True, color="2F5496", size=11),
        "fill": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    },
    "total": {
        "font": Font(bold=True, size=11),
        "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "border": Border(
            top=Side(style="thin", color="2F5496"),
            bottom=Side(style="double", color="2F5496"),
        ),
    },
    "title": {
        "font": Font(bold=True, size=16, color="2F5496"),
        "alignment": Alignment(horizontal="center"),
    },
    "subtitle": {
        "font": Font(size=12, color="4472C4"),
        "alignment": Alignment(horizontal="center"),
    },
    "worst": {
        "font": Font(bold=True, color="FF0000"),
    },
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def apply_style(cell, style_name):
    style = THEME.get(style_name, {})
    if "font" in style:
        cell.font = style["font"]
    if "fill" in style:
        cell.fill = style["fill"]
    if "alignment" in style:
        cell.alignment = style["alignment"]
    if "border" in style:
        cell.border = style["border"]


def write_cell(ws, row, col, cell_data, row_style=None):
    cell = ws.cell(row=row, column=col)

    if cell_data is None:
        cell.value = None
    elif isinstance(cell_data, dict):
        if "f" in cell_data:
            cell.value = cell_data["f"]
        elif "v" in cell_data:
            cell.value = cell_data["v"]
        if "fmt" in cell_data:
            cell.number_format = cell_data["fmt"]
        if "comment" in cell_data:
            cell.comment = Comment(cell_data["comment"], "Bookkeeper")
        if "style" in cell_data:
            apply_style(cell, cell_data["style"])
    else:
        cell.value = cell_data

    if row_style:
        apply_style(cell, row_style)

    cell.border = THIN_BORDER
    return cell


def build_sheet(wb, sheet_spec, is_first=False):
    name = sheet_spec["name"]
    if is_first:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)

    # Column widths
    for i, width in enumerate(sheet_spec.get("columns", []), 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze panes
    freeze = sheet_spec.get("freeze")
    if freeze:
        freeze_row, freeze_col = freeze
        ws.freeze_panes = ws.cell(row=freeze_row + 1, column=freeze_col + 1)

    # Autofilter
    if sheet_spec.get("autofilter"):
        num_cols = len(sheet_spec.get("columns", []))
        num_rows = len([r for r in sheet_spec.get("rows", []) if r is not None])
        if num_cols > 0 and num_rows > 0:
            ws.auto_filter.ref = "A1:{}{}".format(
                get_column_letter(num_cols), num_rows
            )

    # Rows
    row_num = 0
    for row_data in sheet_spec.get("rows", []):
        row_num += 1

        if row_data is None:
            continue

        row_style = row_data.get("style")

        # Merged title/subtitle rows
        if "merge" in row_data:
            merge_cols = row_data["merge"]
            value = row_data.get("value", "")
            cell = ws.cell(row=row_num, column=1, value=value)
            if row_style:
                apply_style(cell, row_style)
            if merge_cols > 1:
                ws.merge_cells(
                    start_row=row_num,
                    start_column=1,
                    end_row=row_num,
                    end_column=merge_cols,
                )
            continue

        # Regular row with cells
        cells = row_data.get("cells", [])
        for col_num, cell_data in enumerate(cells, 1):
            write_cell(ws, row_num, col_num, cell_data, row_style)

    # Conditional formatting
    for cf in sheet_spec.get("conditional_formats", []):
        if cf["type"] == "color_scale":
            rule = ColorScaleRule(
                start_type="min",
                start_color=cf["min_color"].lstrip("#"),
                end_type="max",
                end_color=cf["max_color"].lstrip("#"),
            )
            ws.conditional_formatting.add(cf["range"], rule)

    return ws


def main():
    with open(SPEC_PATH, "r") as f:
        spec = json.load(f)

    wb = openpyxl.Workbook()

    for i, sheet_spec in enumerate(spec["sheets"]):
        build_sheet(wb, sheet_spec, is_first=(i == 0))

    wb.save(OUTPUT_PATH)
    print("[OK] Created: {}".format(OUTPUT_PATH))
    print("[OK] Sheets: {}".format(wb.sheetnames))


if __name__ == "__main__":
    main()
