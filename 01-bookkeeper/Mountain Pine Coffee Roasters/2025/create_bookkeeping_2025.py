import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook with same sheets as 2024 template
wb = openpyxl.Workbook()

# Rename default sheet to Transactions
ws = wb.active
ws.title = 'Transactions'

# Add other sheets
wb.create_sheet('Monthly Summary')
wb.create_sheet('Expense Categories')
wb.create_sheet('Profit & Loss')
wb.create_sheet('Dashboard')

# Set up Transactions sheet headers
headers = ['Date', 'Type', 'Category', 'Vendor/Customer', 'Description', 'Reference', 'Income', 'Expense', 'Balance', 'Tax Deductible', 'Notes']
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)
    ws.cell(row=1, column=col).font = Font(bold=True)

# Transaction data from bank statements
transactions = [
    # January 2025
    ('2025-01-03', 'Expense', 'Cost of Goods Sold', 'Pacific Coffee Importers', 'Green coffee beans', 'Check #1042', None, 1705.00, 'Yes', None),
    ('2025-01-05', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-001', 47.74, None, 'N/A', None),
    ('2025-01-06', 'Expense', 'Internet & Phone', 'Xfinity Business', 'Internet and phone service', 'AUTO', None, 128.99, 'Yes', None),
    ('2025-01-07', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 46.72, 'Yes', None),
    ('2025-01-07', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-002', 39.17, None, 'N/A', None),
    ('2025-01-08', 'Expense', 'Packaging & Supplies', 'PackagingPro Supply', 'Coffee bags, labels, boxes', 'POS', None, 284.34, 'Yes', None),
    ('2025-01-09', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-003', 95.85, None, 'N/A', None),
    ('2025-01-10', 'Expense', 'Utilities', 'Boulder Electric Co.', 'Electric service', 'AUTO', None, 198.45, 'Yes', None),
    ('2025-01-12', 'Expense', 'Shipping & Delivery', 'UPS Store', 'Shipping supplies', 'POS', None, 90.12, 'Yes', None),
    ('2025-01-12', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-004', 53.64, None, 'N/A', None),
    ('2025-01-14', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 51.35, 'Yes', None),
    ('2025-01-14', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-005', 39.68, None, 'N/A', None),
    ('2025-01-15', 'Expense', 'Office Supplies', 'Office Depot', 'Office supplies', 'POS', None, 109.42, 'Yes', None),
    ('2025-01-17', 'Expense', 'Cost of Goods Sold', 'Pacific Coffee Importers', 'Green coffee beans', 'Check #1043', None, 1540.00, 'Yes', None),
    ('2025-01-17', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-006', 39.16, None, 'N/A', None),
    ('2025-01-18', 'Income', 'Income', 'The Daily Grind Cafe', 'Wholesale coffee order', 'INV-2025-001', 386.00, None, 'N/A', None),
    ('2025-01-19', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-007', 60.61, None, 'N/A', None),
    ('2025-01-21', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 43.07, 'Yes', None),
    ('2025-01-22', 'Income', 'Income', 'Sunrise Espresso Bar', 'Wholesale coffee order', 'INV-2025-002', 304.00, None, 'N/A', None),
    ('2025-01-22', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-008', 73.35, None, 'N/A', None),
    ('2025-01-24', 'Expense', 'Equipment & Maintenance', 'Coffee Tech Services', 'Roaster maintenance/repair', 'Check #1044', None, 386.28, 'Yes', None),
    ('2025-01-25', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-009', 43.45, None, 'N/A', None),
    ('2025-01-26', 'Expense', 'Shipping & Delivery', 'UPS Store', 'Shipping supplies', 'POS', None, 75.10, 'Yes', None),
    ('2025-01-28', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 51.84, 'Yes', None),
    ('2025-01-28', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-010', 68.12, None, 'N/A', None),
    # February 2025
    ('2025-02-02', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-011', 53.64, None, 'N/A', None),
    ('2025-02-03', 'Income', 'Income', 'The Daily Grind Cafe', 'Wholesale coffee order', 'Deposit', 482.00, None, 'N/A', None),
    ('2025-02-03', 'Expense', 'Cost of Goods Sold', 'Pacific Coffee Importers', 'Green coffee beans', 'Check #1045', None, 1765.00, 'Yes', None),
    ('2025-02-05', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 49.58, 'Yes', None),
    ('2025-02-05', 'Income', 'Income', 'Sunrise Espresso Bar', 'Wholesale coffee order', 'INV-2025-005', 410.00, None, 'N/A', None),
    ('2025-02-05', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-012', 61.68, None, 'N/A', None),
    ('2025-02-06', 'Expense', 'Internet & Phone', 'Xfinity Business', 'Internet and phone service', 'AUTO', None, 128.99, 'Yes', None),
    ('2025-02-08', 'Expense', 'Packaging & Supplies', 'PackagingPro Supply', 'Coffee bags, labels, boxes', 'POS', None, 531.14, 'Yes', None),
    ('2025-02-08', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-013', 36.46, None, 'N/A', None),
    ('2025-02-10', 'Expense', 'Utilities', 'Boulder Electric Co.', 'Electric service', 'AUTO', None, 215.78, 'Yes', None),
    ('2025-02-10', 'Income', 'Income', 'Mountain View Bistro', 'Wholesale coffee order', 'INV-2025-003', 790.88, None, 'N/A', None),
    ('2025-02-10', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-014', 59.55, None, 'N/A', None),
    ('2025-02-11', 'Expense', 'Shipping & Delivery', 'UPS Store', 'Shipping supplies', 'POS', None, 118.56, 'Yes', None),
    ('2025-02-12', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 50.01, 'Yes', None),
    ('2025-02-13', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-015', 71.33, None, 'N/A', None),
    ('2025-02-14', 'Expense', 'Office Supplies', 'Office Depot', 'Office supplies', 'POS', None, 53.61, 'Yes', None),
    ('2025-02-15', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-016', 45.60, None, 'N/A', None),
    ('2025-02-17', 'Income', 'Income', 'The Daily Grind Cafe', 'Wholesale coffee order', 'Deposit', 548.00, None, 'N/A', None),
    ('2025-02-18', 'Expense', 'Cost of Goods Sold', 'Pacific Coffee Importers', 'Green coffee beans', 'Check #1046', None, 1695.00, 'Yes', None),
    ('2025-02-18', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-017', 58.46, None, 'N/A', None),
    ('2025-02-19', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 47.46, 'Yes', None),
    ('2025-02-19', 'Income', 'Income', 'Sunrise Espresso Bar', 'Wholesale coffee order', 'INV-2025-007', 362.00, None, 'N/A', None),
    ('2025-02-21', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-018', 53.64, None, 'N/A', None),
    ('2025-02-24', 'Expense', 'Shipping & Delivery', 'UPS Store', 'Shipping supplies', 'POS', None, 127.68, 'Yes', None),
    ('2025-02-24', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-019', 86.20, None, 'N/A', None),
    ('2025-02-26', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 51.31, 'Yes', None),
    ('2025-02-27', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-020', 26.29, None, 'N/A', None),
    # March 2025
    ('2025-03-02', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-021', 32.18, None, 'N/A', None),
    ('2025-03-03', 'Expense', 'Cost of Goods Sold', 'Pacific Coffee Importers', 'Green coffee beans', 'Check #1047', None, 1980.00, 'Yes', None),
    ('2025-03-03', 'Income', 'Income', 'The Daily Grind Cafe', 'Wholesale coffee order', 'INV-2025-011', 466.00, None, 'N/A', None),
    ('2025-03-05', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 56.24, 'Yes', None),
    ('2025-03-05', 'Income', 'Income', 'Sunrise Espresso Bar', 'Wholesale coffee order', 'INV-2025-012', 484.00, None, 'N/A', None),
    ('2025-03-05', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-022', 55.24, None, 'N/A', None),
    ('2025-03-06', 'Expense', 'Internet & Phone', 'Xfinity Business', 'Internet and phone service', 'AUTO', None, 128.99, 'Yes', None),
    ('2025-03-07', 'Expense', 'Packaging & Supplies', 'PackagingPro Supply', 'Coffee bags, labels, boxes', 'POS', None, 520.41, 'Yes', None),
    ('2025-03-08', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-023', 63.83, None, 'N/A', None),
    ('2025-03-10', 'Expense', 'Utilities', 'Boulder Electric Co.', 'Electric service', 'AUTO', None, 187.23, 'Yes', None),
    ('2025-03-11', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-024', 53.64, None, 'N/A', None),
    ('2025-03-12', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 51.30, 'Yes', None),
    ('2025-03-12', 'Income', 'Income', 'Mountain View Bistro', 'Wholesale coffee order', 'INV-2025-008', 938.00, None, 'N/A', None),
    ('2025-03-13', 'Expense', 'Shipping & Delivery', 'UPS Store', 'Shipping supplies', 'POS', None, 118.03, 'Yes', None),
    ('2025-03-14', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-025', 64.90, None, 'N/A', None),
    ('2025-03-15', 'Expense', 'Office Supplies', 'Office Depot', 'Office supplies', 'POS', None, 86.88, 'Yes', None),
    ('2025-03-17', 'Expense', 'Cost of Goods Sold', 'Pacific Coffee Importers', 'Green coffee beans', 'Check #1048', None, 2070.00, 'Yes', None),
    ('2025-03-17', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-026', 75.47, None, 'N/A', None),
    ('2025-03-19', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 51.30, 'Yes', None),
    ('2025-03-19', 'Income', 'Income', 'Sunrise Espresso Bar', 'Wholesale coffee order', 'INV-2025-015', 427.00, None, 'N/A', None),
    ('2025-03-20', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-027', 47.74, None, 'N/A', None),
    ('2025-03-21', 'Expense', 'Equipment & Maintenance', 'Coffee Tech Services', 'Roaster maintenance/repair', 'Check #1049', None, 654.53, 'Yes', None),
    ('2025-03-23', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-028', 73.35, None, 'N/A', None),
    ('2025-03-24', 'Expense', 'Shipping & Delivery', 'UPS Store', 'Shipping supplies', 'POS', None, 94.38, 'Yes', None),
    ('2025-03-26', 'Expense', 'Vehicle & Fuel', 'Shell Gas Station', 'Delivery van fuel', 'POS', None, 55.94, 'Yes', None),
    ('2025-03-26', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-029', 37.53, None, 'N/A', None),
    ('2025-03-28', 'Expense', 'Packaging & Supplies', 'PackagingPro Supply', 'Coffee bags, labels, boxes', 'POS', None, 191.00, 'Yes', None),
    ('2025-03-28', 'Income', 'Income', 'Online Sale', 'Online store order', 'RET-2025-030', 53.64, None, 'N/A', None),
]

# Sort by date
transactions.sort(key=lambda x: x[0])

# Calculate running balance starting from Jan 1 beginning balance
balance = 15420.00

for row, txn in enumerate(transactions, 2):
    date_str, txn_type, category, vendor, desc, ref, income, expense, tax_ded, notes = txn
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')

    # Update balance
    if income:
        balance += income
    if expense:
        balance -= expense

    ws.cell(row=row, column=1, value=date_obj)
    ws.cell(row=row, column=2, value=txn_type)
    ws.cell(row=row, column=3, value=category)
    ws.cell(row=row, column=4, value=vendor)
    ws.cell(row=row, column=5, value=desc)
    ws.cell(row=row, column=6, value=ref)
    ws.cell(row=row, column=7, value=income)
    ws.cell(row=row, column=8, value=expense)
    ws.cell(row=row, column=9, value=round(balance, 2))
    ws.cell(row=row, column=10, value=tax_ded)
    ws.cell(row=row, column=11, value=notes)

# Set column widths
col_widths = [12, 10, 22, 25, 30, 15, 12, 12, 12, 14, 20]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Create Expense Categories sheet
ws_cat = wb['Expense Categories']
cat_headers = ['Category', 'Total Q1 2025', '% of Expenses', 'Tax Deductible', 'Tax Form Line']
for col, header in enumerate(cat_headers, 1):
    ws_cat.cell(row=1, column=col, value=header)
    ws_cat.cell(row=1, column=col).font = Font(bold=True)

# Calculate expense totals by category
expense_totals = {}
for txn in transactions:
    if txn[1] == 'Expense' and txn[7]:
        cat = txn[2]
        expense_totals[cat] = expense_totals.get(cat, 0) + txn[7]

total_expenses = sum(expense_totals.values())

categories_data = [
    ('Cost of Goods Sold', expense_totals.get('Cost of Goods Sold', 0), 'Yes', 'Line 4 - COGS'),
    ('Equipment & Maintenance', expense_totals.get('Equipment & Maintenance', 0), 'Yes', 'Line 13 - Depreciation'),
    ('Internet & Phone', expense_totals.get('Internet & Phone', 0), 'Yes', 'Line 25 - Utilities'),
    ('Office Supplies', expense_totals.get('Office Supplies', 0), 'Yes', 'Line 22 - Supplies'),
    ('Packaging & Supplies', expense_totals.get('Packaging & Supplies', 0), 'Yes', 'Line 22 - Supplies'),
    ('Shipping & Delivery', expense_totals.get('Shipping & Delivery', 0), 'Yes', 'Line 27a - Other'),
    ('Utilities', expense_totals.get('Utilities', 0), 'Yes', 'Line 25 - Utilities'),
    ('Vehicle & Fuel', expense_totals.get('Vehicle & Fuel', 0), 'Partial', 'Line 9 - Car/Truck'),
]

for row, (cat, total, tax_ded, tax_line) in enumerate(categories_data, 2):
    ws_cat.cell(row=row, column=1, value=cat)
    ws_cat.cell(row=row, column=2, value=round(total, 2))
    ws_cat.cell(row=row, column=3, value=round(total/total_expenses, 4) if total_expenses > 0 else 0)
    ws_cat.cell(row=row, column=4, value=tax_ded)
    ws_cat.cell(row=row, column=5, value=tax_line)

# Add total row
total_row = len(categories_data) + 2
ws_cat.cell(row=total_row, column=1, value='TOTAL EXPENSES')
ws_cat.cell(row=total_row, column=1).font = Font(bold=True)
ws_cat.cell(row=total_row, column=2, value=round(total_expenses, 2))
ws_cat.cell(row=total_row, column=2).font = Font(bold=True)

# Set column widths for Expense Categories
for i, width in enumerate([25, 15, 15, 15, 20], 1):
    ws_cat.column_dimensions[get_column_letter(i)].width = width

# Create Monthly Summary sheet
ws_sum = wb['Monthly Summary']
sum_headers = ['Month', 'Total Income', 'Total Expenses', 'Net Profit/Loss', 'Beginning Balance', 'Ending Balance']
for col, header in enumerate(sum_headers, 1):
    ws_sum.cell(row=1, column=col, value=header)
    ws_sum.cell(row=1, column=col).font = Font(bold=True)

# Calculate monthly totals
monthly_data = {
    'January': {'income': 0, 'expense': 0, 'begin': 15420.00},
    'February': {'income': 0, 'expense': 0, 'begin': 0},
    'March': {'income': 0, 'expense': 0, 'begin': 0}
}

for txn in transactions:
    date_obj = datetime.strptime(txn[0], '%Y-%m-%d')
    month = date_obj.strftime('%B')
    if txn[6]:  # income
        monthly_data[month]['income'] += txn[6]
    if txn[7]:  # expense
        monthly_data[month]['expense'] += txn[7]

# Calculate ending balances
monthly_data['January']['end'] = monthly_data['January']['begin'] + monthly_data['January']['income'] - monthly_data['January']['expense']
monthly_data['February']['begin'] = monthly_data['January']['end']
monthly_data['February']['end'] = monthly_data['February']['begin'] + monthly_data['February']['income'] - monthly_data['February']['expense']
monthly_data['March']['begin'] = monthly_data['February']['end']
monthly_data['March']['end'] = monthly_data['March']['begin'] + monthly_data['March']['income'] - monthly_data['March']['expense']

row = 2
for month in ['January', 'February', 'March']:
    data = monthly_data[month]
    ws_sum.cell(row=row, column=1, value=month + ' 2025')
    ws_sum.cell(row=row, column=2, value=round(data['income'], 2))
    ws_sum.cell(row=row, column=3, value=round(data['expense'], 2))
    ws_sum.cell(row=row, column=4, value=round(data['income'] - data['expense'], 2))
    ws_sum.cell(row=row, column=5, value=round(data['begin'], 2))
    ws_sum.cell(row=row, column=6, value=round(data['end'], 2))
    row += 1

# Q1 totals
ws_sum.cell(row=5, column=1, value='Q1 2025 TOTAL')
ws_sum.cell(row=5, column=1).font = Font(bold=True)
total_income = sum(d['income'] for d in monthly_data.values())
total_exp = sum(d['expense'] for d in monthly_data.values())
ws_sum.cell(row=5, column=2, value=round(total_income, 2))
ws_sum.cell(row=5, column=3, value=round(total_exp, 2))
ws_sum.cell(row=5, column=4, value=round(total_income - total_exp, 2))
ws_sum.cell(row=5, column=5, value=15420.00)
ws_sum.cell(row=5, column=6, value=round(monthly_data['March']['end'], 2))

# Set column widths for Monthly Summary
for i, width in enumerate([15, 14, 14, 16, 16, 14], 1):
    ws_sum.column_dimensions[get_column_letter(i)].width = width

# Save workbook
wb.save(r'D:\ReposFred\cc-ai-scenarios\01-bookkeeper\Mountain Pine Coffee Roasters\2025\Bookkeeping 2025.xlsx')
print('Spreadsheet created successfully!')
print(f'Total Income Q1: ${total_income:.2f}')
print(f'Total Expenses Q1: ${total_exp:.2f}')
print(f'Net Profit/Loss Q1: ${total_income - total_exp:.2f}')
print(f'Ending Balance: ${monthly_data["March"]["end"]:.2f}')
